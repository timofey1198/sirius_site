# -*- coding: utf-8 -*-
import openpyxl
from random import choice, randint
from PIL import Image, ImageDraw, ImageFont
import sqlite3

def is_access(user, key):
    wb = openpyxl.load_workbook(filename = 'resources/users.xlsx')
    sheet = wb['1']
    users_number = sheet['A1'].value
    for i in range(users_number):
        u = sheet['A%s'%str(i+3)].value
        k = sheet['C%s'%str(i+3)].value
        if u == user and key >= k:
            wb.save('resources/users.xlsx')
            return True
    wb.save('resources/users.xlsx')
    return False

def is_user(user):
    wb = openpyxl.load_workbook(filename = 'resources/users.xlsx')
    sheet = wb['1']
    users_number = sheet['A1'].value
    for i in range(users_number):
        u = sheet['A%s'%str(i+3)].value
        if u == user:
            wb.save('resources/users.xlsx')
            return True
    wb.save('resources/users.xlsx')
    return False

def is_password(user, password):
    wb = openpyxl.load_workbook(filename = 'resources/users.xlsx')
    sheet = wb['1']
    users_number = sheet['A1'].value    
    for i in range(users_number):
        u = sheet['A%s'%str(i+3)].value
        if u == user:
            p = sheet['B%s'%str(i+3)].value
            if p == password:
                wb.save('resources/users.xlsx')
                return True
    wb.save('resources/users.xlsx')
    return False

def new_user(user, password):
    wb = openpyxl.load_workbook(filename = 'resources/users.xlsx')
    sheet = wb['1']
    users_number = sheet['A1'].value
    sheet['A%s' %str(users_number + 3)] = user
    sheet['B%s' %str(users_number + 3)] = password
    sheet['C%s' %str(users_number + 3)] = 3
    sheet['A1'] = users_number + 1
    wb.save('resources/users.xlsx')


def captcha():
    key = ''.join([choice('QWERTYUIOPLKJHGFDSAZXCVBNM1234567890') for i in range(5)])
    img = Image.new('RGB', (180,60), 0xffffff )
    draw = ImageDraw.Draw(img)
    for i in range(60):
        draw.line( [(randint(0,200),randint(0,60)), 
                    (randint(0,200),randint(0,60))], 
                   randint(0, 0xffffff), 1)
    font = ImageFont.truetype('resources/fonts/7fonts.ru_Funky09-Bold.ttf', 60)
    draw.text( (0,0), key, 0, font)
    img.save('resources/img/captcha/%s.jpg' %str(key), 'JPEG')
    return key

def name_parser(name):
    first_split = list(name.split('_'))
    num = first_split[0]
    title = list(first_split[1].split('.'))[0]
    return (title, num)

 



# Создание таблицы
#cursor.execute("""CREATE TABLE users_access
                  #(id integer primary key, login text, password text,  
                   #access_num integer)
               #""")


              #cursor.execute("""INSERT INTO users_access
                            #VALUES (1, 'timostar', 'Upiter98', 0)"""
                         #)
          
          #

def get_last_nubmer(table):
    conn = sqlite3.connect("resources/data.db")
    cursor = conn.cursor()
    sql = "SELECT MAX([id]) FROM %s"%table
    cursor.execute(sql)
    return cursor.fetchall()[0][0]

def new_article(title, body, author):
    number = get_last_nubmer('news')
    if number == None:
        number = 0
    number += 1
    conn = sqlite3.connect("resources/data.db")
    cursor = conn.cursor()
    filename = str(number)
    cursor.execute("""INSERT INTO news
                   VALUES (?, ?, ?, ?)
                   """, [str(number), title, filename, author])
    conn.commit()
    f = open('pages/news/%s.txt'%filename, 'w+', encoding='utf-8')
    f.write(body)
    f.close()

def get_articles_info(number):
    conn = sqlite3.connect("resources/data.db")
    cursor = conn.cursor()
    cursor.execute("""SELECT * FROM news WHERE id=?
                   """, [("%i"%number)])
    return cursor.fetchall()[0]

def delete_article(number):
    conn = sqlite3.connect("resources/data.db")
    cursor = conn.cursor()
    cursor.execute("DELETE FROM news WHERE id=?", [("%i"%number)])
    conn.commit()

if __name__ == "__main__":
    print(get_last_nubmer('news'))
    #print(is_access('timostar', 0))